import customtkinter as ctk
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import time
import os
import re
from threading import Thread
from datetime import datetime, timedelta

# 엑셀 저장용
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side

# --- GUI 설정 ---
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")


def clean_filename(text):
    return re.sub(r'[\\/*?:"<>|]', "", text)


class CafeUrlGroup(ctk.CTkFrame):
    def __init__(self, master, on_delete, placeholder_text=""):
        super().__init__(master, fg_color="transparent")
        self.on_delete = on_delete
        self.pack(fill="x", pady=2, padx=5)
        self.entry_url = ctk.CTkEntry(self, placeholder_text=placeholder_text, height=32)
        self.entry_url.pack(side="left", padx=(5, 5), pady=2, expand=True, fill="x")
        self.btn_del = ctk.CTkButton(self, text="✕", width=35, height=32, fg_color="#e74c3c", hover_color="#c0392b",
                                     command=self.delete_self)
        self.btn_del.pack(side="right", padx=5)

    def delete_self(self):
        self.on_delete(self)
        self.destroy()


class CategoryGroup(ctk.CTkFrame):
    def __init__(self, master, on_delete):
        super().__init__(master, fg_color="transparent")
        self.on_delete = on_delete
        self.pack(fill="x", pady=2, padx=5)
        self.entry_cat = ctk.CTkEntry(self, placeholder_text="카테고리", width=100, height=32)
        self.entry_cat.pack(side="left", padx=5, pady=2)
        self.option_target = ctk.CTkOptionMenu(self, values=["글+댓글", "제목만", "작성자", "댓글"], width=110, height=32)
        self.option_target.set("글+댓글")
        self.option_target.pack(side="left", padx=5, pady=2)
        self.entry_kws = ctk.CTkEntry(self, placeholder_text="키워드 (쉼표 구분)", height=32)
        self.entry_kws.pack(side="left", padx=5, pady=2, expand=True, fill="x")
        self.btn_del = ctk.CTkButton(self, text="✕", width=35, height=32, fg_color="#e74c3c", hover_color="#c0392b",
                                     command=self.delete_self)
        self.btn_del.pack(side="right", padx=5)

    def delete_self(self):
        self.on_delete(self)
        self.destroy()


class CafeDetailedScraper(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Naver Cafe Pro Scraper v3.9 (Monthly Summary Enhanced)")
        self.geometry("1650x950")

        self.cafe_groups = []
        self.category_groups = []
        self.main_container = ctk.CTkFrame(self, fg_color="transparent")
        self.main_container.pack(fill="both", expand=True, padx=15, pady=15)

        self.left_panel = ctk.CTkFrame(self.main_container, width=850)
        self.left_panel.pack(side="left", fill="both", expand=False, padx=(0, 10))
        self.left_panel.pack_propagate(False)

        ctk.CTkLabel(self.left_panel, text="⚙️ SCRAPER SETTINGS", font=("Pretendard", 22, "bold"),
                     text_color="#3498db").pack(pady=20)
        self.btn_start = ctk.CTkButton(self.left_panel, text="🚀 수집 시작", command=self.start_multi_threads,
                                       fg_color="#2ecc71", hover_color="#27ae60", height=65,
                                       font=("Pretendard", 18, "bold"))
        self.btn_start.pack(side="bottom", pady=25, padx=20, fill="x")

        self.scroll_area = ctk.CTkScrollableFrame(self.left_panel, fg_color="transparent")
        self.scroll_area.pack(side="top", fill="both", expand=True, padx=5, pady=5)

        self.box_login = self.create_box(self.scroll_area, "🔑 로그인 설정")
        self.entry_id = ctk.CTkEntry(self.box_login, placeholder_text="아이디", width=250)
        self.entry_id.pack(side="left", padx=25, pady=15, expand=True)
        self.entry_pw = ctk.CTkEntry(self.box_login, placeholder_text="비밀번호", width=250, show="*")
        self.entry_pw.pack(side="left", padx=25, pady=15, expand=True)

        self.box_cafe = self.create_box(self.scroll_area, "🌐 카페 URL")
        self.cafe_container = ctk.CTkFrame(self.box_cafe, fg_color="transparent")
        self.cafe_container.pack(fill="x", padx=5)
        ctk.CTkButton(self.box_cafe, text="+ URL 추가", fg_color="#34495e", height=30, command=self.add_cafe).pack(
            pady=10)
        self.add_cafe(placeholder_text="https://cafe.naver.com/...")

        self.box_date = self.create_box(self.scroll_area, "📅 분석 기간 (월별 자동 구분)")
        self.date_inner = ctk.CTkFrame(self.box_date, fg_color="transparent")
        self.date_inner.pack(pady=10)
        self.entry_start = ctk.CTkEntry(self.date_inner, placeholder_text="시작일", width=180)
        self.entry_start.insert(0, "2026-01-01")
        self.entry_start.pack(side="left", padx=10)
        ctk.CTkLabel(self.date_inner, text="~", font=("Pretendard", 18)).pack(side="left")
        self.entry_end = ctk.CTkEntry(self.date_inner, placeholder_text="종료일", width=180)
        self.entry_end.insert(0, "2026-03-02")
        self.entry_end.pack(side="left", padx=10)

        self.box_kw = self.create_box(self.scroll_area, "🔍 분류 및 필터")
        self.group_container = ctk.CTkFrame(self.box_kw, fg_color="transparent")
        self.group_container.pack(fill="x", padx=5)
        ctk.CTkButton(self.box_kw, text="+ 카테고리 추가", fg_color="#34495e", height=30, command=self.add_category).pack(
            pady=10)
        self.add_category()

        self.right_panel = ctk.CTkFrame(self.main_container)
        self.right_panel.pack(side="right", fill="both", expand=True)
        self.log_text = ctk.CTkTextbox(self.right_panel, font=("Consolas", 14), border_width=1)
        self.log_text.pack(fill="both", expand=True, padx=15, pady=15)

    def create_box(self, master, title):
        box = ctk.CTkFrame(master, border_width=2, border_color="#3B3B3B")
        box.pack(pady=10, padx=10, fill="x")
        ctk.CTkLabel(box, text=title, font=("Pretendard", 13, "bold")).pack(pady=5)
        return box

    def add_cafe(self, placeholder_text="https://cafe.naver.com/..."):
        group = CafeUrlGroup(self.cafe_container, self.remove_cafe, placeholder_text=placeholder_text)
        self.cafe_groups.append(group)

    def remove_cafe(self, group):
        if group in self.cafe_groups: self.cafe_groups.remove(group)

    def add_category(self):
        group = CategoryGroup(self.group_container, self.remove_category)
        self.category_groups.append(group)

    def remove_category(self, group):
        if group in self.category_groups: self.category_groups.remove(group)

    def log(self, message, cafe_id="SYSTEM"):
        self.log_text.insert("end", f"[{time.strftime('%H:%M:%S')}][{cafe_id}] {message}\n")
        self.log_text.see("end")

    def start_multi_threads(self):
        urls = [g.entry_url.get().strip() for g in self.cafe_groups if g.entry_url.get().strip()]
        if not urls: return
        self.btn_start.configure(state="disabled", text="⏳ 수집 중...")
        for target_url in urls:
            Thread(target=self.run_scraper_for_cafe, args=(target_url,), daemon=True).start()
        self.after(5000, lambda: self.btn_start.configure(state="normal", text="🚀 수집 시작"))

    def get_monthly_ranges(self, start_str, end_str):
        start = datetime.strptime(start_str, "%Y-%m-%d")
        end = datetime.strptime(end_str, "%Y-%m-%d")
        months = []
        curr = start
        while curr <= end:
            m_start = curr
            if curr.month == 12:
                next_m = curr.replace(year=curr.year + 1, month=1, day=1)
            else:
                next_m = curr.replace(month=curr.month + 1, day=1)
            m_end = next_m - timedelta(days=1)
            if m_end > end: m_end = end
            months.append({"label": m_start.strftime("%Y년 %m월"), "start": m_start.strftime("%Y-%m-%d"),
                           "end": m_end.strftime("%Y-%m-%d")})
            curr = next_m
        return months

    def run_scraper_for_cafe(self, target_url):
        raw_cafe_id = target_url.rstrip('/').split('/')[-1]
        cafe_id = clean_filename(raw_cafe_id)
        month_ranges = self.get_monthly_ranges(self.entry_start.get().strip(), self.entry_end.get().strip())
        user_id, user_pw = self.entry_id.get().strip(), self.entry_pw.get().strip()

        tasks = []
        for g in self.category_groups:
            cat, target, kws = g.entry_cat.get().strip(), g.option_target.get(), [k.strip() for k in
                                                                                  g.entry_kws.get().split(",") if
                                                                                  k.strip()]
            if cat and kws: tasks.append({"category": cat, "target": target, "keywords": kws})

        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=Options())
        wait = WebDriverWait(driver, 15)
        all_data = []

        try:
            if user_id and user_pw:
                driver.get("https://nid.naver.com/nidlogin.login")
                time.sleep(1);
                driver.execute_script(
                    f"document.getElementsByName('id')[0].value = '{user_id}'; document.getElementsByName('pw')[0].value = '{user_pw}';")
                driver.find_element(By.ID, "log.login").click();
                time.sleep(2)

            for m_range in month_ranges:
                for task in tasks:
                    for keyword in task["keywords"]:
                        self.log(f"🔄 [{keyword}] ({m_range['label']}) 검색", cafe_id)
                        driver.get(target_url);
                        time.sleep(2)
                        if "/menus/" not in target_url:
                            try:
                                driver.switch_to.default_content()
                                wait.until(
                                    EC.element_to_be_clickable((By.XPATH, "//a[contains(text(), '전체글보기')]"))).click()
                                time.sleep(1)
                            except:
                                pass
                        try:
                            driver.switch_to.default_content()
                            wait.until(EC.frame_to_be_available_and_switch_to_it((By.ID, "cafe_main")))
                        except:
                            pass

                        self.perform_optimized_search_logic(driver, wait, keyword, m_range["start"], m_range["end"],
                                                            task["target"], cafe_id)

                        post_links = []
                        current_page = 1
                        while True:
                            time.sleep(2.5)
                            rows = driver.find_elements(By.CSS_SELECTOR, "tr")
                            p_cnt = 0
                            for row in rows:
                                try:
                                    link = row.find_element(By.CSS_SELECTOR, "a.article").get_attribute("href")
                                    if link not in post_links: post_links.append(link); p_cnt += 1
                                except:
                                    continue
                            if p_cnt == 0 or not self.move_to_next_page(driver, current_page): break
                            current_page += 1

                        self.log(f"📝 {len(post_links)}개 글 수집 중...", cafe_id)
                        for url in post_links:
                            driver.get(url);
                            time.sleep(1.8)
                            try:
                                driver.switch_to.default_content()
                                wait.until(EC.frame_to_be_available_and_switch_to_it((By.ID, "cafe_main")))
                                p_title = driver.find_element(By.CSS_SELECTOR, "h3.title_text").text
                                p_author = driver.find_element(By.CSS_SELECTOR, ".nickname").text
                                p_date = driver.find_element(By.CSS_SELECTOR, ".date").text
                                p_content = driver.find_element(By.CSS_SELECTOR, ".article_viewer").text
                                c_list = []
                                for itm in driver.find_elements(By.CSS_SELECTOR, ".CommentItem"):
                                    try:
                                        c_list.append({
                                            "author": itm.find_element(By.CSS_SELECTOR, ".comment_nickname").text,
                                            "content": itm.find_element(By.CSS_SELECTOR, ".comment_text_view").text,
                                            "date": itm.find_element(By.CSS_SELECTOR, ".comment_info_date").text
                                        })
                                    except:
                                        continue
                                all_data.append({"cat": task["category"], "target": task["target"], "kw": keyword,
                                                 "month": m_range["label"], "url": url, "title": p_title,
                                                 "author": p_author, "date": p_date, "content": p_content,
                                                 "comments": c_list})
                            except:
                                continue

            if all_data:
                self.save_styled_flat_excel(all_data, cafe_id)
                self.save_monthly_summary_excel(all_data, cafe_id)  # ★ 강화된 요약 보고서 호출

        except Exception as e:
            self.log(f"❌ 오류: {str(e)}", cafe_id)
        finally:
            driver.quit();
            self.log(f"🏁 작업 완료", cafe_id)

    def perform_optimized_search_logic(self, driver, wait, keyword, start, end, search_target, cafe_id):
        try:
            search_input = wait.until(
                EC.presence_of_element_located((By.CSS_SELECTOR, ".search_input_area input, #topLayerQueryInput")))
            search_input.clear();
            search_input.send_keys(keyword);
            search_input.send_keys(Keys.ENTER)
            time.sleep(2.5)
            if search_target != "글+댓글":
                try:
                    target_btn = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "div.target button")))
                    driver.execute_script("arguments[0].click();", target_btn);
                    time.sleep(1)
                    target_option = driver.find_element(By.XPATH,
                                                        f"//div[contains(@class, 'target')]//*[contains(text(), '{search_target}')]")
                    driver.execute_script("arguments[0].click();", target_option);
                    time.sleep(1)
                except:
                    pass
            detail_btn = wait.until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, ".btn_detail, .period .button, .MoreArrowButton")))
            driver.execute_script("arguments[0].click();", detail_btn);
            time.sleep(1)

            def react_type(name, val):
                el = driver.find_element(By.NAME, name)
                el.click();
                el.send_keys(Keys.CONTROL + "a");
                el.send_keys(Keys.BACKSPACE)
                for c in val: el.send_keys(c)
                el.send_keys(Keys.TAB)

            react_type("min", start);
            react_type("max", end)
            driver.execute_script("arguments[0].click();",
                                  driver.find_element(By.CSS_SELECTOR, "button.BaseButton--skinGray, .btn_apply"))
            time.sleep(1);
            driver.execute_script("arguments[0].click();", driver.find_element(By.CSS_SELECTOR, "button.btn_search"))
            time.sleep(1)
        except Exception as e:
            self.log(f"⚠️ 검색 설정 실패: {e}", cafe_id)

    # --- [강화된 월별 요약 보고서 엑셀 저장] ---
    def save_monthly_summary_excel(self, data, cafe_name):
        try:
            filename = f"{cafe_name}_월별요약_{datetime.now().strftime('%m%d_%H%M')}.xlsx"
            full_path = os.path.join(os.path.join(os.path.expanduser("~"), "Downloads"), filename)
            wb = Workbook();
            ws = wb.active
            ws.title = "월별 요약"

            # 데이터 집계 (월, 카테고리, 조건, 키워드별)
            counts = {}
            for item in data:
                key = (item["month"], item["cat"], item["target"], item["kw"])
                counts[key] = counts.get(key, 0) + 1

            headers = ["월", "카테고리", "검색조건", "키워드", "게시글 수"]
            ws.append(headers)

            h_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
            month_colors = [
                PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
                PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
                PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
                PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
            ]
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                            bottom=Side(style='thin'))
            h_font, bold_font = Font(color="FFFFFF", bold=True), Font(bold=True)

            for i, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=i)
                cell.fill, cell.font, cell.alignment, cell.border = h_fill, h_font, Alignment(
                    horizontal="center"), border

            current_row = 2
            month_start_row = 2
            cat_start_row = 2
            last_month = None
            last_cat = None
            color_idx = -1

            # 집계된 데이터 정렬하여 쓰기
            sorted_keys = sorted(counts.keys())
            for key in sorted_keys:
                month, cat, target, kw = key
                count = counts[key]

                if month != last_month:
                    if last_month is not None:
                        # 월 병합
                        ws.merge_cells(start_row=month_start_row, start_column=1, end_row=current_row - 1, end_column=1)
                        ws.cell(row=month_start_row, column=1).font = bold_font
                        # 마지막 카테고리 병합
                        ws.merge_cells(start_row=cat_start_row, start_column=2, end_row=current_row - 1, end_column=2)

                    month_start_row = current_row
                    cat_start_row = current_row
                    last_month = month
                    last_cat = cat
                    color_idx += 1

                if cat != last_cat:
                    # 카테고리 병합
                    ws.merge_cells(start_row=cat_start_row, start_column=2, end_row=current_row - 1, end_column=2)
                    cat_start_row = current_row
                    last_cat = cat

                ws.append([month, cat, target, kw, count])
                fill = month_colors[color_idx % len(month_colors)]
                for col in range(1, 6):
                    c = ws.cell(row=current_row, column=col)
                    c.fill, c.border = fill, border
                    c.alignment = Alignment(horizontal="center", vertical="center")
                current_row += 1

            # 마지막 섹션 병합
            ws.merge_cells(start_row=month_start_row, start_column=1, end_row=current_row - 1, end_column=1)
            ws.cell(row=month_start_row, column=1).font = bold_font
            ws.merge_cells(start_row=cat_start_row, start_column=2, end_row=current_row - 1, end_column=2)

            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['D'].width = 30
            wb.save(full_path)
            self.log(f"✅ 월별 요약 보고서 완료", cafe_name)
        except Exception as e:
            self.log(f"❌ 요약 엑셀 오류: {e}", cafe_name)

    def save_styled_flat_excel(self, data, cafe_name):
        try:
            filename = f"{cafe_name}_상세결과_{datetime.now().strftime('%m%d_%H%M')}.xlsx"
            full_path = os.path.join(os.path.join(os.path.expanduser("~"), "Downloads"), filename)
            wb = Workbook();
            ws = wb.active
            header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
            month_colors = [PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
                            PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
                            PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
                            PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")]
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                 bottom=Side(style='thin'))
            white_font, bold_font = Font(color="FFFFFF", bold=True), Font(bold=True)

            max_c = max([len(d["comments"]) for d in data]) if data else 0
            headers = ["카테고리", "검색조건", "키워드", "월", "URL", "제목", "작성자", "작성일", "본문내용"]
            for i in range(1, max_c + 1): headers += [f"댓글{i}작성자", f"댓글{i}내용", f"댓글{i}작성일"]
            ws.append(headers)
            for i in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=i)
                cell.fill, cell.font, cell.alignment, cell.border = header_fill, white_font, Alignment(
                    horizontal="center"), thin_border

            current_row, month_start_row, last_month, color_idx = 2, 2, None, -1
            for idx, item in enumerate(data):
                if item["month"] != last_month:
                    if last_month is not None:
                        ws.merge_cells(start_row=month_start_row, start_column=4, end_row=current_row - 1, end_column=4)
                        ws.cell(row=month_start_row, column=4).font, ws.cell(row=month_start_row,
                                                                             column=4).alignment = bold_font, Alignment(
                            horizontal="center", vertical="center")
                    month_start_row, last_month, color_idx = current_row, item["month"], color_idx + 1
                row_vals = [item["cat"], item["target"], item["kw"], item["month"], item["url"], item["title"],
                            item["author"], item["date"], item["content"]]
                for c in item["comments"]: row_vals += [c["author"], c["content"], c["date"]]
                ws.append(row_vals)
                fill = month_colors[color_idx % len(month_colors)]
                for col in range(1, len(row_vals) + 1):
                    c = ws.cell(row=current_row, column=col)
                    c.fill, c.border, c.alignment = fill, thin_border, Alignment(wrap_text=True, vertical="top")
                current_row += 1
            ws.merge_cells(start_row=month_start_row, start_column=4, end_row=current_row - 1, end_column=4)
            ws.cell(row=month_start_row, column=4).font, ws.cell(row=month_start_row,
                                                                 column=4).alignment = bold_font, Alignment(
                horizontal="center", vertical="center")
            ws.column_dimensions['F'].width, ws.column_dimensions['I'].width = 35, 60
            wb.save(full_path);
            self.log(f"✅ 상세 데이터 완료", cafe_name)
        except Exception as e:
            self.log(f"❌ 상세 엑셀 오류: {str(e)}", cafe_name)

    def move_to_next_page(self, driver, current_page):
        try:
            page_numbers = driver.find_elements(By.CSS_SELECTOR, ".Pagination a, .Pagination button.number")
            for page in page_numbers:
                if page.text.strip() == str(current_page + 1):
                    driver.execute_script("arguments[0].click();", page);
                    return True
            next_arrow = driver.find_elements(By.CSS_SELECTOR, ".Pagination button.type_next, .pgR")
            if next_arrow and next_arrow[0].is_enabled():
                driver.execute_script("arguments[0].click();", next_arrow[0]);
                return True
        except:
            pass
        return False


if __name__ == "__main__":
    app = CafeDetailedScraper()
    app.mainloop()