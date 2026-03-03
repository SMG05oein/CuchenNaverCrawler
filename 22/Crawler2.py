import time
import os
import re
from datetime import datetime, timedelta
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side


class CafeScraperEngine:
    def __init__(self, log_callback):
        self.log_callback = log_callback

    def log(self, message, cafe_id="SYSTEM"):
        self.log_callback(message, cafe_id)

    def clean_filename(self, text):
        return re.sub(r'[\\/*?:"<>|]', "", text)

    def get_monthly_ranges(self, start_str, end_str):
        start = datetime.strptime(start_str, "%Y-%m-%d")
        end = datetime.strptime(end_str, "%Y-%m-%d")
        months = []
        curr = start
        while curr <= end:
            m_start = curr
            if curr.month == 12:
                next_m = curr.replace(year=curr.year + 1, month=1, day=1)
            else:
                next_m = curr.replace(month=curr.month + 1, day=1)
            m_end = next_m - timedelta(days=1)
            if m_end > end: m_end = end
            months.append({"label": m_start.strftime("%Y년 %m월"), "start": m_start.strftime("%Y-%m-%d"),
                           "end": m_end.strftime("%Y-%m-%d")})
            curr = next_m
        return months

    def run_scraper_for_cafe(self, target_url, params):
        cafe_id = self.clean_filename(target_url.rstrip('/').split('/')[-1])
        month_ranges = self.get_monthly_ranges(params['start_date'], params['end_date'])

        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=Options())
        wait = WebDriverWait(driver, 15)
        all_data = []

        try:
            if params['id'] and params['pw']:
                driver.get("https://nid.naver.com/nidlogin.login")
                time.sleep(1)
                driver.execute_script(
                    f"document.getElementsByName('id')[0].value = '{params['id']}'; document.getElementsByName('pw')[0].value = '{params['pw']}';")
                driver.find_element(By.ID, "log.login").click()
                time.sleep(2)

            for m_range in month_ranges:
                for task in params['tasks']:
                    for keyword in task["keywords"]:
                        self.log(f"🔄 [{keyword}] ({m_range['label']}) 검색", cafe_id)
                        driver.get(target_url);
                        time.sleep(2)

                        if "/menus/" not in target_url:
                            try:
                                driver.switch_to.default_content()
                                wait.until(
                                    EC.element_to_be_clickable((By.XPATH, "//a[contains(text(), '전체글보기')]"))).click()
                                time.sleep(1)
                            except:
                                pass

                        try:
                            driver.switch_to.default_content()
                            wait.until(EC.frame_to_be_available_and_switch_to_it((By.ID, "cafe_main")))
                        except:
                            pass

                        self.perform_optimized_search_logic(driver, wait, keyword, m_range["start"], m_range["end"],
                                                            task["target"], cafe_id)

                        post_links = []
                        current_page = 1
                        while True:
                            time.sleep(2.5)
                            rows = driver.find_elements(By.CSS_SELECTOR, "tr")
                            p_cnt = 0
                            for row in rows:
                                try:
                                    link = row.find_element(By.CSS_SELECTOR, "a.article").get_attribute("href")
                                    if link not in post_links: post_links.append(link); p_cnt += 1
                                except:
                                    continue
                            if p_cnt == 0 or not self.move_to_next_page(driver, current_page): break
                            current_page += 1

                        self.log(f"📝 {len(post_links)}개 글 수집 중...", cafe_id)
                        for url in post_links:
                            driver.get(url);
                            time.sleep(1.8)
                            try:
                                driver.switch_to.default_content()
                                wait.until(EC.frame_to_be_available_and_switch_to_it((By.ID, "cafe_main")))
                                p_title = driver.find_element(By.CSS_SELECTOR, "h3.title_text").text
                                p_author = driver.find_element(By.CSS_SELECTOR, ".nickname").text
                                p_date = driver.find_element(By.CSS_SELECTOR, ".date").text
                                p_content = driver.find_element(By.CSS_SELECTOR, ".article_viewer").text
                                c_list = []
                                for itm in driver.find_elements(By.CSS_SELECTOR, ".CommentItem"):
                                    try:
                                        c_list.append({
                                            "author": itm.find_element(By.CSS_SELECTOR, ".comment_nickname").text,
                                            "content": itm.find_element(By.CSS_SELECTOR, ".comment_text_view").text,
                                            "date": itm.find_element(By.CSS_SELECTOR, ".comment_info_date").text
                                        })
                                    except:
                                        continue
                                all_data.append({"cat": task["category"], "target": task["target"], "kw": keyword,
                                                 "month": m_range["label"], "url": url, "title": p_title,
                                                 "author": p_author, "date": p_date, "content": p_content,
                                                 "comments": c_list})
                            except:
                                continue

            if all_data:
                self.save_styled_flat_excel(all_data, cafe_id)
                self.save_monthly_summary_excel(all_data, cafe_id)

        except Exception as e:
            self.log(f"❌ 오류: {str(e)}", cafe_id)
        finally:
            driver.quit()
            self.log(f"🏁 작업 완료", cafe_id)

    def perform_optimized_search_logic(self, driver, wait, keyword, start, end, search_target, cafe_id):
        try:
            search_input = wait.until(
                EC.presence_of_element_located((By.CSS_SELECTOR, ".search_input_area input, #topLayerQueryInput")))
            search_input.clear();
            search_input.send_keys(keyword);
            search_input.send_keys(Keys.ENTER)
            time.sleep(2.5)
            if search_target != "글+댓글":
                try:
                    target_btn = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "div.target button")))
                    driver.execute_script("arguments[0].click();", target_btn);
                    time.sleep(1)
                    target_option = driver.find_element(By.XPATH,
                                                        f"//div[contains(@class, 'target')]//*[contains(text(), '{search_target}')]")
                    driver.execute_script("arguments[0].click();", target_option);
                    time.sleep(1)
                except:
                    pass
            detail_btn = wait.until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, ".btn_detail, .period .button, .MoreArrowButton")))
            driver.execute_script("arguments[0].click();", detail_btn);
            time.sleep(1)

            def react_type(name, val):
                el = driver.find_element(By.NAME, name)
                el.click();
                el.send_keys(Keys.CONTROL + "a");
                el.send_keys(Keys.BACKSPACE)
                for c in val: el.send_keys(c)
                el.send_keys(Keys.TAB)

            react_type("min", start);
            react_type("max", end)
            driver.execute_script("arguments[0].click();",
                                  driver.find_element(By.CSS_SELECTOR, "button.BaseButton--skinGray, .btn_apply"))
            time.sleep(1);
            driver.execute_script("arguments[0].click();", driver.find_element(By.CSS_SELECTOR, "button.btn_search"))
            time.sleep(1)
        except Exception as e:
            self.log(f"⚠️ 검색 설정 실패: {e}", cafe_id)

    def save_monthly_summary_excel(self, data, cafe_name):
        try:
            filename = f"{cafe_name}_월별요약_{datetime.now().strftime('%m%d_%H%M')}.xlsx"
            full_path = os.path.join(os.path.join(os.path.expanduser("~"), "Downloads"), filename)
            wb = Workbook();
            ws = wb.active
            ws.title = "월별 요약"
            counts = {}
            for item in data:
                key = (item["month"], item["cat"], item["target"], item["kw"])
                counts[key] = counts.get(key, 0) + 1
            headers = ["월", "카테고리", "검색조건", "키워드", "게시글 수"]
            ws.append(headers)
            h_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
            month_colors = [PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
                            PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
                            PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
                            PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")]
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                            bottom=Side(style='thin'))
            h_font, bold_font = Font(color="FFFFFF", bold=True), Font(bold=True)
            for i, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=i)
                cell.fill, cell.font, cell.alignment, cell.border = h_fill, h_font, Alignment(
                    horizontal="center"), border
            current_row, month_start_row, cat_start_row, last_month, last_cat, color_idx = 2, 2, 2, None, None, -1
            sorted_keys = sorted(counts.keys())
            for key in sorted_keys:
                month, cat, target, kw = key
                count = counts[key]
                if month != last_month:
                    if last_month is not None:
                        ws.merge_cells(start_row=month_start_row, start_column=1, end_row=current_row - 1, end_column=1)
                        ws.cell(row=month_start_row, column=1).font = bold_font
                        ws.merge_cells(start_row=cat_start_row, start_column=2, end_row=current_row - 1, end_column=2)
                    month_start_row, cat_start_row, last_month, last_cat, color_idx = current_row, current_row, month, cat, color_idx + 1
                if cat != last_cat:
                    ws.merge_cells(start_row=cat_start_row, start_column=2, end_row=current_row - 1, end_column=2)
                    cat_start_row, last_cat = current_row, cat
                ws.append([month, cat, target, kw, count])
                fill = month_colors[color_idx % len(month_colors)]
                for col in range(1, 6):
                    c = ws.cell(row=current_row, column=col)
                    c.fill, c.border, c.alignment = fill, border, Alignment(horizontal="center", vertical="center")
                current_row += 1
            ws.merge_cells(start_row=month_start_row, start_column=1, end_row=current_row - 1, end_column=1)
            ws.cell(row=month_start_row, column=1).font = bold_font
            ws.merge_cells(start_row=cat_start_row, start_column=2, end_row=current_row - 1, end_column=2)
            ws.column_dimensions['A'].width, ws.column_dimensions['B'].width, ws.column_dimensions[
                'D'].width = 15, 20, 30
            wb.save(full_path)
        except Exception as e:
            self.log(f"❌ 요약 엑셀 오류: {e}", cafe_name)

    def save_styled_flat_excel(self, data, cafe_name):
        try:
            filename = f"{cafe_name}_상세결과_{datetime.now().strftime('%m%d_%H%M')}.xlsx"
            full_path = os.path.join(os.path.join(os.path.expanduser("~"), "Downloads"), filename)
            wb = Workbook();
            ws = wb.active
            header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
            month_colors = [PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
                            PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
                            PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
                            PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")]
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                 bottom=Side(style='thin'))
            white_font, bold_font = Font(color="FFFFFF", bold=True), Font(bold=True)
            max_c = max([len(d["comments"]) for d in data]) if data else 0
            headers = ["카테고리", "검색조건", "키워드", "월", "URL", "제목", "작성자", "작성일", "본문내용"]
            for i in range(1, max_c + 1): headers += [f"댓글{i}작성자", f"댓글{i}내용", f"댓글{i}작성일"]
            ws.append(headers)
            for i in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=i)
                cell.fill, cell.font, cell.alignment, cell.border = header_fill, white_font, Alignment(
                    horizontal="center"), thin_border
            current_row, month_start_row, last_month, color_idx = 2, 2, None, -1
            for idx, item in enumerate(data):
                if item["month"] != last_month:
                    if last_month is not None:
                        ws.merge_cells(start_row=month_start_row, start_column=4, end_row=current_row - 1, end_column=4)
                        ws.cell(row=month_start_row, column=4).font, ws.cell(row=month_start_row,
                                                                             column=4).alignment = bold_font, Alignment(
                            horizontal="center", vertical="center")
                    month_start_row, last_month, color_idx = current_row, item["month"], color_idx + 1
                row_vals = [item["cat"], item["target"], item["kw"], item["month"], item["url"], item["title"],
                            item["author"], item["date"], item["content"]]
                for c in item["comments"]: row_vals += [c["author"], c["content"], c["date"]]
                ws.append(row_vals)
                fill = month_colors[color_idx % len(month_colors)]
                for col in range(1, len(row_vals) + 1):
                    c = ws.cell(row=current_row, column=col)
                    c.fill, c.border, c.alignment = fill, thin_border, Alignment(wrap_text=True, vertical="top")
                current_row += 1
            ws.merge_cells(start_row=month_start_row, start_column=4, end_row=current_row - 1, end_column=4)
            ws.cell(row=month_start_row, column=4).font, ws.cell(row=month_start_row,
                                                                 column=4).alignment = bold_font, Alignment(
                horizontal="center", vertical="center")
            ws.column_dimensions['F'].width, ws.column_dimensions['I'].width = 35, 60
            wb.save(full_path)
        except Exception as e:
            self.log(f"❌ 상세 엑셀 오류: {str(e)}", cafe_name)

    def move_to_next_page(self, driver, current_page):
        try:
            page_numbers = driver.find_elements(By.CSS_SELECTOR, ".Pagination a, .Pagination button.number")
            for page in page_numbers:
                if page.text.strip() == str(current_page + 1):
                    driver.execute_script("arguments[0].click();", page);
                    return True
            next_arrow = driver.find_elements(By.CSS_SELECTOR, ".Pagination button.type_next, .pgR")
            if next_arrow and next_arrow[0].is_enabled():
                driver.execute_script("arguments[0].click();", next_arrow[0]);
                return True
        except:
            pass
        return False